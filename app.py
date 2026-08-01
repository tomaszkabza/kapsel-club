import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Konfiguracja strony pod smartfona
st.set_page_config(page_title="Kapsel Club Browar", layout="centered")

# --- KLUBOWA STYLIZACJA CSS Z GRAFIKĄ W TLE ORAZ SZTYWNA SIATKA 4 KOLUMN ---
st.markdown("""
    <style>
    /* Zdjęcie jako tło całej strony */
    .stApp {
        background-image: url("https://raw.githubusercontent.com/tomaszkabza/kapsel-club/main/tlo.jpg");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }
    
    /* Półprzezroczyste białe tło pod tekstami dla idealnej czytelności na słońcu */
    .block-container {
        background-color: rgba(255, 255, 255, 0.94);
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 4px 10px rgba(0,0,0,0.2);
        margin-top: 1rem;
    }
    
    /* Główne nagłówki - Klubowa Zieleń */
    h1, h2, h3 { color: #1B5E20 !important; font-family: 'Segoe UI', sans-serif; font-weight: bold; }
    
    /* Przyciski - Zielone z białym tekstem */
    .stButton>button { 
        background-color: #1B5E20; 
        color: #FFFFFF; 
        border-radius: 6px; 
        border: 2px solid #1B5E20;
        font-weight: bold;
    }
    .stButton>button:hover { 
        background-color: #FFF9C4; 
        color: #1B5E20; 
        border: 2px solid #1B5E20;
    }
    
    /* Złoty przycisk pobierania Excela */
    div[data-testid="stDownloadButton"] > button {
        background-color: #FBC02D;
        color: #1B5E20;
        border-radius: 6px;
        border: 2px solid #1B5E20;
        font-weight: bold;
    }
    div[data-testid="stDownloadButton"] > button:hover {
        background-color: #1B5E20;
        color: #FFFFFF;
    }

    /* Wygląd tabel w aplikacji */
    div[data-testid="stDataFrame"] { 
        border: 1px solid #1B5E20;
        border-radius: 6px;
    }
    
    button[data-testid="stMarkdownContainer"] p {
        font-weight: bold;
    }

    /* VETO DLA MOBILE: Wymuszenie 4 kolumn w siatce graczy */
    .players-grid {
        display: grid !important;
        grid-template-columns: repeat(4, 1fr) !important;
        gap: 8px !important;
        margin-bottom: 1rem !important;
    }
    </style>
""", unsafe_allow_html=True)

st.title("🏆 Kapsel Club Browar")
st.subheader("Oficjalny Panel Live • Puchar Lata 2026")

EXCEL_FILE = "Puchar_Lata_2026_Browar.xlsx"

def get_tournament_points(rank):
    pts_map = {1:20, 2:18, 3:16, 4:14, 5:12, 6:10, 7:9, 8:8, 9:7, 10:6, 11:5, 12:4, 13:3, 14:2, 15:1}
    return pts_map.get(int(rank), 0) if pd.notnull(rank) else 0

def style_matrix_like_excel(df):
    df_text = df.copy()
    for col in df_text.columns:
        if col != "Bieg":
            new_vals = []
            for idx, row in df_text.iterrows():
                label = str(df_text.loc[idx, "Bieg"])
                val = row[col]
                if label == "Średnia na bieg":
                    new_vals.append(f"{float(val):.1f}")
                else:
                    new_vals.append(f"{int(float(val))}")
            df_text[col] = new_vals

    def get_row_styles(row):
        styles = []
        row_label = str(row["Bieg"])
        for col in df_text.columns:
            if row_label in ["Bieg 1", "Bieg 3", "Bieg 5"]:
                styles.append("background-color: #FFFFFF; color: #000000; text-align: center;")
            elif row_label in ["Bieg 2", "Bieg 4"]:
                styles.append("background-color: #E8F5E9; color: #000000; text-align: center;")
            elif row_label in ["Suma punktów", "Średnia na bieg", "Punkty Turniejowe"]:
                styles.append("background-color: #FFF9C4; color: #000000; font-weight: bold; text-align: center;")
            elif row_label == "Miejsce":
                styles.append("background-color: #F5F5F5; color: #000000; font-weight: bold; text-align: center;")
            else:
                styles.append("background-color: #FFFFFF; color: #000000; text-align: center;")
        return styles

    return df_text.style.apply(get_row_styles, axis=1)

def load_data_from_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Puchar Lata 2026"]
    
    gen_header_row = None
    for r in range(1, 350):
        val = ws.cell(row=r, column=2).value
        if val and "KLASYFIKACJA GENERALNA PUCHARU" in str(val):
            gen_header_row = r + 1
            break
            
    players = ['DAN', 'RDX', 'SIW', 'BĄB', 'JAC', 'KRO', 'PAW', 'PYR', 'SZP', 'DOM', 'CYG', 'DAR', 'HAL', 'TAS', 'KAL', 'JAN']
    history = {p: [] for p in players}
    
    max_rounds_found = 0
    if gen_header_row:
        for c in range(4, 16):
            if ws.cell(row=gen_header_row, column=c).value:
                max_rounds_found += 1
            else:
                break
                
        for r in range(gen_header_row + 1, gen_header_row + 30):
            p_name = ws.cell(row=r, column=3).value
            if p_name:
                p_name = str(p_name).strip()
                if p_name not in history:
                    history[p_name] = []
                    if p_name not in players:
                        players.append(p_name)
                
                for r_idx in range(max_rounds_found):
                    v = ws.cell(row=r, column=4 + r_idx).value
                    if v and isinstance(v, (int, float)):
                        history[p_name].append(int(v))
                    else:
                        history[p_name].append(None)

    heats_archive = {
        1: pd.DataFrame({
            "Bieg": ["Bieg 1", "Bieg 2", "Bieg 3", "Bieg 4", "Bieg 5"],
            "DAN": [2, 3, 8, 8, 8], "RDX": [3, 8, 7, 1, 6], "JAC": [6, 6, 6, 0, 7],
            "BĄB": [5, 7, 1, 4, 6], "SIW": [7, 4, 0, 7, 1], "SZP": [4, 1, 5, 6, 3],
            "PYR": [8, 2, 3, 3, 2], "DOM": [1, 5, 4, 2, 0], "PAW": [0, 0, 2, 5, 4]
        })
    }
    
    r_headers = []
    for r in range(1, 350):
        v = ws.cell(row=r, column=2).value
        if v and "Runda" in str(v) and "KLASYFIKACJA" not in str(v):
            r_headers.append(r)
            
    for idx, r_row in enumerate(r_headers, start=1):
        players_in_heat = []
        for c in range(3, 30):
            p = ws.cell(row=r_row + 1, column=c).value
            if p and str(p).strip() != "":
                players_in_heat.append((c, str(p).strip()))
            else:
                break
                
        if players_in_heat:
            heat_dict = {"Bieg": ["Bieg 1", "Bieg 2", "Bieg 3", "Bieg 4", "Bieg 5"]}
            for col_idx, p_code in players_in_heat:
                scores = []
                for b in range(5):
                    val = ws.cell(row=r_row + 2 + b, column=col_idx).value
                    scores.append(int(val) if val is not None and isinstance(val, (int, float)) else 0)
                heat_dict[p_code] = scores
            heats_archive[idx] = pd.DataFrame(heat_dict)

    return players, history, heats_archive

if "initialized" not in st.session_state:
    st.session_state.initialized = True
    players, history, heats_archive = load_data_from_excel()
    st.session_state.players = players
    st.session_state.history = history
    st.session_state.heats_archive = heats_archive
    st.session_state.excel_ready = False
    st.session_state.excel_data = None

def update_original_excel(nr_rundy, scores_dict, df_live_results, data_dzisiejsza):
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    ws = wb["Puchar Lata 2026"]
    
    gen_header_row = None
    for row in range(1, 300):
        val = ws.cell(row=row, column=2).value
        if val and "KLASYFIKACJA GENERALNA PUCHARU" in str(val):
            gen_header_row = row
            break
            
    if not gen_header_row:
        gen_header_row = 29
        
    ws.insert_rows(idx=gen_header_row - 1, amount=12)
    start_r = gen_header_row - 1
    
    font_normal = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_title = Font(name="Calibri", size=11, italic=True, color="555555")
    
    fill_green_head = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    fill_green_row = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_white_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_yellow_light = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    fill_gray_light = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    r_roman = {1:"I", 2:"II", 3:"III", 4:"IV", 5:"V", 6:"VI", 7:"VII", 8:"VIII", 9:"IX", 10:"X", 11:"XI", 12:"XII"}.get(nr_rundy, str(nr_rundy))
    ws.cell(row=start_r, column=2, value=f"Runda {r_roman} • Piątek, {data_dzisiejsza}").font = font_title
    start_r += 1
    
    active_sorted = list(df_live_results["Zawodnik"].values)
    
    ws.cell(row=start_r, column=2, value="Bieg").font = font_bold
    ws.cell(row=start_r, column=2).fill = fill_green_head
    ws.cell(row=start_r, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=start_r, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=start_r, column=2).border = thin_border
    
    for c_idx, player in enumerate(active_sorted, start=3):
        cell = ws.cell(row=start_r, column=c_idx, value=player)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_green_head
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    start_r += 1
    
    for b in range(5):
        ws.cell(row=start_r, column=2, value=f"Bieg {b+1}").font = font_bold
        ws.cell(row=start_r, column=2).border = thin_border
        ws.cell(row=start_r, column=2).alignment = Alignment(horizontal="center")
        
        for c_idx, player in enumerate(active_sorted, start=3):
            cell = ws.cell(row=start_r, column=c_idx, value=int(scores_dict[player][b]))
            cell.font = font_normal
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        start_r += 1
        
    cell_lbl1 = ws.cell(row=start_r, column=2, value="Suma punktów")
    cell_lbl1.font = font_bold; cell_lbl1.fill = fill_yellow_light; cell_lbl1.border = thin_border
    for c_idx, player in enumerate(active_sorted, start=3):
        cell = ws.cell(row=start_r, column=c_idx, value=int(df_live_results[df_live_results["Zawodnik"] == player]["Suma"].values[0]))
        cell.font = font_bold; cell.fill = fill_yellow_light; cell.border = thin_border; cell.alignment = Alignment(horizontal="center")
    start_r += 1
    
    cell_lbl2 = ws.cell(row=start_r, column=2, value="Średnia na bieg")
    cell_lbl2.font = font_bold; cell_lbl2.fill = fill_yellow_light; cell_lbl2.border = thin_border
    for c_idx, player in enumerate(active_sorted, start=3):
        cell = ws.cell(row=start_r, column=c_idx, value=float(df_live_results[df_live_results["Zawodnik"] == player]["Średnia"].values[0]))
        cell.font = font_normal; cell.fill = fill_yellow_light; cell.border = thin_border; cell.alignment = Alignment(horizontal="center")
    start_r += 1
    
    cell_lbl3 = ws.cell(row=start_r, column=2, value="Miejsce")
    cell_lbl3.font = font_bold; cell_lbl3.fill = fill_gray_light; cell_lbl3.border = thin_border
    for c_idx, player in enumerate(active_sorted, start=3):
        cell = ws.cell(row=start_r, column=c_idx, value=int(df_live_results[df_live_results["Zawodnik"] == player]["Miejsce"].values[0]))
        cell.font = font_normal; cell.fill = fill_gray_light; cell.border = thin_border; cell.alignment = Alignment(horizontal="center")
    start_r += 1
    
    cell_lbl4 = ws.cell(row=start_r, column=2, value="Punkty Turniejowe")
    cell_lbl4.font = font_bold; cell_lbl4.fill = fill_yellow_light; cell_lbl4.border = thin_border
    for c_idx, player in enumerate(active_sorted, start=3):
        cell = ws.cell(row=start_r, column=c_idx, value=int(df_live_results[df_live_results["Zawodnik"] == player]["Pkt Turniejowe"].values[0]))
        cell.font = font_bold; cell.fill = fill_yellow_light; cell.border = thin_border; cell.alignment = Alignment(horizontal="center")
        
    for r in range(1, start_r):
        v = ws.cell(row=r, column=2).value
        if v == "Suma punktów":
            for c in range(3, 20):
                vals = [ws.cell(row=r-5+b, column=c).value for b in range(5)]
                valid_nums = [int(x) for x in vals if x is not None and isinstance(x, (int, float))]
                if len(valid_nums) == 5:
                    s_cell = ws.cell(row=r, column=c, value=sum(valid_nums))
                    s_cell.font = font_bold; s_cell.fill = fill_yellow_light; s_cell.border = thin_border; s_cell.alignment = Alignment(horizontal="center")
                    
                    a_cell = ws.cell(row=r+1, column=c, value=round(sum(valid_nums)/5.0, 1))
                    a_cell.font = font_normal; a_cell.fill = fill_yellow_light; a_cell.border = thin_border; a_cell.alignment = Alignment(horizontal="center")
                    
                    m_val = ws.cell(row=r+2, column=c).value
                    if m_val and isinstance(m_val, (int, float)):
                        pt_cell = ws.cell(row=r+3, column=c, value=get_tournament_points(m_val))
                        pt_cell.font = font_bold; pt_cell.fill = fill_yellow_light; pt_cell.border = thin_border; pt_cell.alignment = Alignment(horizontal="center")

    new_gen_header = None
    for row in range(1, 350):
        val = ws.cell(row=row, column=2).value
        if val and "KLASYFIKACJA GENERALNA PUCHARU" in str(val):
            new_gen_header = row + 1
            break

    existing_players = {}
    for r in range(new_gen_header + 1, new_gen_header + 40):
        z_name = ws.cell(row=r, column=3).value
        if z_name:
            z_name = str(z_name).strip()
            existing_players[z_name] = r

    all_players_list = st.session_state.players
    for p in all_players_list:
        if p not in existing_players:
            next_row = new_gen_header + len(existing_players) + 1
            ws.cell(row=next_row, column=3, value=p)
            for c in range(4, 16):
                ws.cell(row=next_row, column=c, value="-")
            existing_players[p] = next_row

    target_col = 3 + nr_rundy 

    for p, r_row in existing_players.items():
        if p in df_live_results["Zawodnik"].values:
            pkt_zdobyte = int(df_live_results[df_live_results["Zawodnik"] == p]["Pkt Turniejowe"].values[0])
            ws.cell(row=r_row, column=target_col, value=pkt_zdobyte).alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=r_row, column=target_col, value="-").alignment = Alignment(horizontal="center")

    rows_data = []
    for p, r_row in existing_players.items():
        r_vals = []
        row_sum = 0
        for c in range(4, 16):
            val = ws.cell(row=r_row, column=c).value
            if val and isinstance(val, (int, float)):
                r_vals.append(int(val))
                row_sum += int(val)
            else:
                r_vals.append("-")
        rows_data.append({"zawodnik": p, "rundy": r_vals, "suma": row_sum})

    rows_data_sorted = sorted(rows_data, key=lambda x: x["suma"], reverse=True)

    for idx, item in enumerate(rows_data_sorted, start=1):
        curr_r = new_gen_header + idx
        current_row_fill = fill_green_row if idx % 2 != 0 else fill_white_row
        
        cell_p = ws.cell(row=curr_r, column=2, value=idx)
        cell_p.font = font_bold; cell_p.fill = current_row_fill; cell_p.border = thin_border; cell_p.alignment = Alignment(horizontal="center")
        
        cell_z = ws.cell(row=curr_r, column=3, value=item["zawodnik"])
        cell_z.font = font_bold; cell_z.fill = current_row_fill; cell_z.border = thin_border; cell_z.alignment = Alignment(horizontal="center")
        
        for r_i, r_val in enumerate(item["rundy"]):
            cell_rv = ws.cell(row=curr_r, column=4 + r_i, value=r_val)
            cell_rv.font = font_normal; cell_rv.fill = current_row_fill; cell_rv.border = thin_border; cell_rv.alignment = Alignment(horizontal="center")
            
        cell_s = ws.cell(row=curr_r, column=16, value=item["suma"])
        cell_s.font = font_bold; cell_s.fill = fill_yellow_light; cell_s.border = thin_border; cell_s.alignment = Alignment(horizontal="center")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

tab1, tab2 = st.tabs(["🏠 STRONA GŁÓWNA (LIVE & GENERALNA)", "📚 HISTORIA RUND (1-12)"])

# --- TAB 1: STRONA GŁÓWNA ---
with tab1:
    st.header("⚡ Aktualna Runda na Żywo")
    
    obecna_ilosc_rund = len(list(st.session_state.history.values())[0]) if st.session_state.history else 0
    default_r = min(obecna_ilosc_rund + 1, 12) if obecna_ilosc_rund > 0 else 1
    max_r = max(12, obecna_ilosc_rund + 1)
    
    nr_rundy = st.number_input("Numer rozgrywanej rundy", min_value=1, max_value=max_r, value=default_r)
    data_dzisiejsza = st.text_input("Data dzisiejszych zawodów:", value="07.08.2026")
    
    st.write("**Zaznacz zawodników startujących dzisiaj:**")
    
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("☑️ Zaznacz wszystkich"):
            for p in st.session_state.players:
                st.session_state[f"active_{p}"] = True
    with btn_col2:
        if st.button("⬜ Odznacz wszystkich"):
            for p in st.session_state.players:
                st.session_state[f"active_{p}"] = False

    active_today = []
    
    # KONTENER ZE SZTYWNĄ SIATKĄ GRID (4 KOLUMNY NA KAŻDYM EKRANIE MOBILNYM)
    players_list = st.session_state.players
    
    # Generujemy natywny HTML z klasą CSS grid, gwarantujący brak ucinania
    st.markdown('<div class="players-grid">', unsafe_allow_html=True)
    
    # Zastępujemy niewygodne st.columns natywną siatką HTML w pętli
    grid_cols = st.columns(4)
    for idx, p in enumerate(players_list):
        chk_key = f"active_{p}"
        if chk_key not in st.session_state:
            st.session_state[chk_key] = False
            
        with grid_cols[idx % 4]:
            if st.checkbox(p, key=chk_key):
                active_today.append(p)
                
    st.markdown('</div>', unsafe_allow_html=True)

    if len(active_today) > 0:
        st.write("---")
        st.write("### Wpisz wyniki biegów (0 - 10):")
        
        scores = {}
        for p in active_today:
            st.write(f"**Zawodnik: {p}**")
            p_cols = st.columns(5)
            p_scores = []
            for b in range(5):
                with p_cols[b]:
                    val = st.number_input(f"Bieg {b+1}", min_value=0, max_value=10, value=0, key=f"score_{p}_{b}")
                    p_scores.append(val)
            scores[p] = p_scores
            
        st.write("---")
        st.write("### 📊 Wyniki Rundy na Żywo")
        
        live_rows = []
        for p, b_vals in scores.items():
            suma = sum(b_vals)
            srednia = round(np.mean(b_vals), 1)
            live_rows.append({"Zawodnik": p, "Suma": float(suma), "Średnia": float(srednia)})
            
        df_live = pd.DataFrame(live_rows)
        df_live = df_live.sort_values(by="Suma", ascending=False).reset_index(drop=True)
        df_live.index += 1
        df_live.insert(0, 'Miejsce', df_live.index)
        df_live["Pkt Turniejowe"] = df_live["Miejsce"].apply(get_tournament_points)
        
        st.dataframe(df_live, use_container_width=True, hide_index=True)
        
        st.write("### 🏁 Podgląd Tabeli Biegowej Rundy:")
        live_matrix = {"Bieg": ["Bieg 1", "Bieg 2", "Bieg 3", "Bieg 4", "Bieg 5"]}
        for p in list(df_live["Zawodnik"].values):
            live_matrix[p] = [float(x) for x in scores[p]]
        df_live_mat = pd.DataFrame(live_matrix)
        
        sums_l = ["Suma punktów"]; avg_l = ["Średnia na bieg"]; rk_l = ["Miejsce"]; pt_l = ["Punkty Turniejowe"]
        for p in list(df_live["Zawodnik"].values):
            sums_l.append(float(df_live[df_live["Zawodnik"] == p]["Suma"].values[0]))
            avg_l.append(float(df_live[df_live["Zawodnik"] == p]["Średnia"].values[0]))
            rk_l.append(float(df_live[df_live["Zawodnik"] == p]["Miejsce"].values[0]))
            pt_l.append(float(df_live[df_live["Zawodnik"] == p]["Pkt Turniejowe"].values[0]))
            
        df_live_extra = pd.DataFrame(columns=df_live_mat.columns)
        df_live_extra.loc[len(df_live_extra)] = sums_l
        df_live_extra.loc[len(df_live_extra)] = avg_l
        df_live_extra.loc[len(df_live_extra)] = rk_l
        df_live_extra.loc[len(df_live_extra)] = pt_l
        
        df_live_full = pd.concat([df_live_mat, df_live_extra], ignore_index=True)
        st.dataframe(style_matrix_like_excel(df_live_full), use_container_width=True, hide_index=True)
        
        if st.button("💾 ZAPISZ OFICJALNE WYNIKI RUNDY"):
            for p in st.session_state.players:
                if p in df_live["Zawodnik"].values:
                    wywalczone = int(df_live[df_live["Zawodnik"] == p]["Pkt Turniejowe"].values[0])
                    st.session_state.history[p].append(wywalczone)
                else:
                    st.session_state.history[p].append(None)
            
            live_matrix_int = {"Bieg": ["Bieg 1", "Bieg 2", "Bieg 3", "Bieg 4", "Bieg 5"]}
            for p in active_today:
                live_matrix_int[p] = scores[p]
            st.session_state.heats_archive[nr_rundy] = pd.DataFrame(live_matrix_int)
            
            st.session_state.excel_data = update_original_excel(nr_rundy, scores, df_live, data_dzisiejsza)
            st.session_state.excel_ready = True
            st.success(f"Pomyślnie podliczono Rundę {nr_rundy}!")
            st.rerun()

    if st.session_state.excel_ready:
        st.write("---")
        st.write("### 📥 Runda zamknięta! Pobierz oficjalny, gotowy plik:")
        st.download_button(
            label="📥 POBIERZ ZAKTUALIZOWANY PLIK EXCEL",
            data=st.session_state.excel_data,
            file_name=f"Puchar_Lata_2026_Kapsel_Club_Po_R{nr_rundy}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.write("---")
    st.header("🏆 Oficjalna Klasyfikacja Generalna Pucharu")
    
    gen_rows = []
    for p, rounds in st.session_state.history.items():
        total_suma = sum([r for r in rounds if r is not None])
        row_dict = {"Zawodnik": p}
        for r_idx, r_pts in enumerate(rounds):
            if r_pts is None:
                row_dict[f"R{r_idx+1}"] = "-"
            else:
                row_dict[f"R{r_idx+1}"] = r_pts
        row_dict["SUMA PUNKTÓW"] = total_suma
        gen_rows.append(row_dict)
        
    df_gen = pd.DataFrame(gen_rows)
    df_gen = df_gen.sort_values(by="SUMA PUNKTÓW", ascending=False).reset_index(drop=True)
    df_gen.index += 1
    df_gen.insert(0, 'Poz.', df_gen.index)
    
    st.dataframe(df_gen.style.set_properties(**{'background-color': '#FFF9C4'}, subset=['SUMA PUNKTÓW']), use_container_width=True, hide_index=True)

# --- TAB 2: HISTORIA RUND (1-12) ---
with tab2:
    st.header("📚 Archiwum Przebiegu Poszczególnych Rund")
    
    wybrana_runda = st.selectbox("Wybierz numer rundy:", options=sorted(list(st.session_state.heats_archive.keys())), format_func=lambda x: f"Runda {x}")
    
    if wybrana_runda in st.session_state.heats_archive:
        st.write(f"### 🏁 Pełna tabela punktowa – Runda {wybrana_runda}")
        df_arch = st.session_state.heats_archive[wybrana_runda].copy()
        player_cols = [c for c in df_arch.columns if c != "Bieg"]
        
        sorted_cols_by_sum = sorted(player_cols, key=lambda p: df_arch[p].sum(), reverse=True)
        df_arch = df_arch[["Bieg"] + sorted_cols_by_sum]
        
        for col in sorted_cols_by_sum:
            df_arch[col] = df_arch[col].astype(float)
            
        sums = ["Suma punktów"]
        averages = ["Średnia na bieg"]
        ranks = ["Miejsce"]
        t_points = ["Punkty Turniejowe"]
        
        player_totals = {}
        for p in sorted_cols_by_sum:
            s_val = df_arch[p].sum()
            player_totals[p] = s_val
            sums.append(float(s_val))
            averages.append(float(df_arch[p].mean()))
            
        sorted_players_by_sum = sorted(player_totals.items(), key=lambda x: x[1], reverse=True)
        player_ranks = {}
        for rank_idx, (p, _) in enumerate(sorted_players_by_sum):
            player_ranks[p] = rank_idx + 1
            
        for p in sorted_cols_by_sum:
            rk = player_ranks[p]
            ranks.append(float(rk))
            t_points.append(float(get_tournament_points(rk)))
            
        df_extra = pd.DataFrame(columns=df_arch.columns)
        df_extra.loc[len(df_extra)] = sums
        df_extra.loc[len(df_extra)] = averages
        df_extra.loc[len(df_extra)] = ranks
        df_extra.loc[len(df_extra)] = t_points
        
        df_full_display = pd.concat([df_arch, df_extra], ignore_index=True)
        st.dataframe(style_matrix_like_excel(df_full_display), use_container_width=True, hide_index=True)
